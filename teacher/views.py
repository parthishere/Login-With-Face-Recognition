import io
import openpyxl
import pandas
from django.shortcuts import render, redirect, reverse
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from recognizer.models import SessionAttendanceModel, UserProfile, LectrueModel
from login_details.models import LoginDetails
from .forms import SessionForm, TeacherUpdateForm, LectureForm, StudentBulkForm
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from .tasks import create_student




# Create your views here.

def profile_view(request):
    context = {}
    try:
        user = request.user

        user_profile = user.user_profile

        if user.is_teacher:
            context['teacher'] = user_profile
            context['user_profile'] = user_profile
            context['sessions'] = user_profile.change_website_objects.all()
            context['lectures'] = user_profile.lectures.all()
            context['login_objects'] = user_profile.login_details_with_teacher.all(
            ).order_by('-login_date').order_by('-login_time')
        else:
            return redirect('recognizer:login')
    except Exception as e:
        print(e)
        return redirect('recognizer:logout-cnf')

    return render(request, 'teacher/index.html', context=context)


def profile_list_view(request):
    context = {}
    try:
        user = request.user
        user_profile = user.user_profile
        teacher = user_profile if user.is_teacher else None
        context['teacher'] = teacher
        context['user_profile'] = user_profile

        # teachers_user_profiles = TeacherProfileModel.objects.all().values('user')
        students = UserProfile.objects.select_related("user").filter(
            college=teacher.college, branch=teacher.branch).exclude(user__is_teacher=True)
        context['objects'] = students
        context['is_student'] = "Student"
    except Exception as e:
        print(e)
        return redirect('recognizer:logout-cnf')
    return render(request, 'teacher/students-list.html', context=context)


@user_passes_test(lambda u: u.is_staff)
def teacher_profile_list_view(request):
    context = {}
    try:
        user = request.user
        teacher = request.user.user_profile
        context['teacher'] = teacher if teacher.is_teacher else None
        context['user_profile'] = teacher

        if user.is_teacher:
            teachers = UserProfile.objects.filter(
                college=teacher.college).values('user')
            context['is_teacher'] = "Teacher"
            context['objects'] = teachers
        else:
            return redirect("recognizer:login")
    except Exception as e:
        print(e)
        return redirect('recognizer:logout-cnf')

    return render(request, 'teacher/teacher-list.html', context=context)


def teacher_profile_update_view(request):
    edit_form = None
    instance = None
    try:
        instance = request.user.user_profile
    except:
        instance = None

    edit_form = TeacherUpdateForm(request.POST or None, instance=instance)

    context = {
        'form': edit_form,
    }
    if instance.user == request.user:
        if request.POST:
            if edit_form.is_valid:
                user = edit_form.save()

                messages.success(request, "Teacher Profile Edited Sucsessfuly")
                context = {
                    'form': edit_form,
                }
                return redirect("teacher:dashboard")
            else:
                context = {
                    'form': edit_form,
                }
                messages.error(request, "Somthing is wrong ..")
                return redirect("teacher:dashboard")

    return render(request, 'teacher/update-teacher-profile.html', context=context)


def update_ips(request):
    try:
        user = request.user
        teacher = user.user_profile
        if request.POST and user.is_teacher:
            print("ohk")
            ip1 = request.POST['ip1']
            teacher.ip_address1 = ip1
            teacher.save()
            messages.success(request, "IP Updated Sucsessfuly")
            return redirect("recognizer:home")

    except:
        return redirect("recognizer:login")


def lecture_list_view(request):
    context = {}
    try:
        user = request.user
        teacher = UserProfile.objects.select_related(
            "user", "college", "branch").prefetch_related("lectures").get(user=user)
        teacher = user.user_profile
        context['teacher'] = teacher if user.is_teacher else None
        context['user_profile'] = teacher

        if user.is_teacher:
            lectures = teacher.lectures.all()
            context['objects'] = lectures

    except Exception as e:
        print(e)
        return redirect('recognizer:logout-cnf')

    return render(request, 'teacher/lectures.html', context=context)


@login_required(login_url='recognizer:login')
def lec_detail_view(request, pk=None):
    context = {}
    print(pk)
    lecture = LectrueModel.objects.prefetch_realated(
        "change_website_objects_lecture", "lecture_login_details", "accepted_user", "requested_user", "requested_user__branch_name").get(pk=pk)
    print(lecture)
    context['lecture'] = lecture
    context['user'] = request.user
    return render(request, 'teacher/lectures_detail.html', context=context)


def add_lecture(request):
    context = {}
    form = LectureForm(request.POST or None)
    context['form'] = form
    user = request.user
    user_profile = user.user_profile
    if request.POST and form.is_valid and user.is_teacher:
        teacher_profile = user_profile if user.is_teacher else None
        if teacher_profile:
            obj = form.save(commit=False)
            obj.teacher = teacher_profile
            obj.save()
            context['form'] = form
            messages.success(request, "Lecture Added Succsessfully")
            return redirect('teacher:lec')
        else:
            return redirect("recognizer:login")

    return render(request, 'teacher/update-teacher-profile.html', context=context)


def search_student(request):
    try:
        context = {}
        query = request.GET['q']
        user_profile = request.user.user_profile
        qs = UserProfile.objects.select_related("user").filter(
            Q(user__username__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(user__enrollment_number__icontains=query) |
            Q(phone_number__icontains=query)
        ).filter(college=user_profile.college)
        context['userprofile'] = qs.filter(user__is_teacher=False)

        teacher = qs.filter(user__is_teacher=True)
        context['teacher'] = teacher

        context['lectures'] = LectrueModel.objects.filter(Q(lecture_name__icontains=query) | Q(
            code__exact=query)).filter(college=user_profile.college)
        context['objects'] = qs
    except Exception as e:
        print(e)
        return redirect("recognizer:login")
    return render(request, "teacher/search-list.html", context)


def search_lectures(request):
    query = request.GET.get('q')
    return (request, "", {})


def session_detail(request, pk):
    context = {}
    obj = SessionAttendanceModel.objects.select_related(
        "lecture", "teacher").prefetch_related("atendees", "requested_users").get(pk=pk)
    context['session'] = obj
    return render(request, "teacher/session-detail.html", context)


def delete_session(request, pk=None):
    session = SessionAttendanceModel.objects.get(pk=pk)
    if request.user.user_profile is session.teacher:
        session.delete()
        return redirect('teacher:lec')
    else:
        return redirect("recognizer:login")


def delete_lecture(request, pk=None):
    lecture = LectrueModel.objects.get(pk=pk)
    if request.user.user_profile is lecture.teacher and lecture:
        lecture.delete()
        return redirect('teacher:lec')
    else:
        return redirect("recognizer:login")


def update_session(request, pk=None):
    context = {}
    user = request.user
    user_profile = UserProfile.objects.prefetch_related(
        "lectures").get(user=user)
    session = SessionAttendanceModel.objects.get(pk=pk)
    users = UserProfile.objects.filter(college=user_profile.college)

    form = SessionForm(request.POST or None, instance=session,
                       user_profile=users, lectures=user_profile.lectures.all())
    context["form"] = form
    if form.is_valid() and session.teacher == user_profile and user.is_teacher:
        form.save()
        context["form"] = form
        return redirect(reverse("teacher:session-detail", pk=session.pk))
    return render(request, "teacher/update-teacher-profile.html", context)


def delete_attendance(request, pk=None):
    att = LoginDetails.objects.get(pk=pk)
    if att.teacher.user is request.user:
        print('ok')
        att.delete()
        return redirect('teacher:dashboard')
    else:
        return redirect("recognizer:login")


def update_lecture(request, pk=None):
    context = {}
    lec = LectrueModel.objects.get(pk=pk)
    form = LectureForm(request.POST or None, instance=lec)
    context['form'] = form

    if lec.teacher == request.user.user_profile and form.is_valid():
        instance = form.save()
        messages.success(request, "Lecture Updated Succsessfully")
        return redirect("teacher:lec")
    return render(request, 'teacher/update-teacher-profile.html', context)


def reset_confirm_view(request, pk=None):
    context = {}
    lecture = LectrueModel.objects.get(pk=pk)
    context['view'] = 'Reset Lecture'
    context['msg'] = f"Reset Lecture {lecture.lecture_name}??"
    context['lecture'] = lecture
    return render(request, 'teacher/reset-cnf.html', context=context)


def reset_attendance_of_lecture(request, pk=None):
    # lecture = LectrueModel.objects.get(pk=pk)
    teacher = request.user
    teacher_profile = UserProfile.objects.prefetch_related(
        "lectures", "login_details_with_teacher", "change_website_objects").get(user=request.user)
    lecture = teacher_profile.lectures.get(pk=pk)
    if lecture.teacher == teacher_profile and teacher.is_teacher:
        teacher_profile.login_details_with_teacher.filter(
            lecture=lecture, teacher=teacher_profile).delete()
        teacher_profile.change_website_objects.filter(
            lecture=lecture, teacher=teacher_profile).delete()
        return redirect("teacher:dashboard")


def lec_detail(request, pk=None):
    context = {}
    lec = LectrueModel.objects.select_related("teacher").prefetch_related(
        "requested_user", "accepted_user", "change_website_objects_lecture", "lecture_login_details").get(pk=pk)

    context['lecture'] = lec
    return render(request, "teacher/lecture_detail.html", context)


def send_request(request, pk):
    lecture = LectrueModel.objects.get(pk=pk)
    user = request.user
    user_profile = user.user_profile
    if not user.is_teacher:
        lecture.requested_user.add(user_profile)
        lecture.save()
    return redirect(reverse("teacher:lec-detail", kwargs={'pk': lecture.pk}))


def accept_request(request, user_id, lec_id):
    lecture = LectrueModel.objects.get(id=lec_id)
    user_p = UserProfile.objects.get(pk=user_id)
    user = request.user
    user_profile = user.user_profile
    teacher_profile = user_profile if user.is_teacher else None
    if teacher_profile == lecture.teacher:
        if user_p in lecture.accepted_user.all().iterator():
            if user_p in lecture.requested_user.all().iterator():
                lecture.requested_user.remove(user_p)
        else:
            lecture.accepted_user.add(user_p)
            lecture.requested_user.remove(user_p)
    return redirect(reverse("teacher:lec-detail", kwargs={'pk': lecture.pk}))


def decline_request(request, user_id, lec_id):
    lecture = LectrueModel.objects.get(id=lec_id)
    user_p = UserProfile.objects.get(pk=user_id)
    user = request.user
    user_profile = user.user_profile
    teacher_profile = user_profile if user.is_teacher else None
    if teacher_profile == lecture.teacher:

        if user_p in lecture.accepted_user.all().iterator():
            lecture.accepted_user.remove(user_p)
            try:
                lecture.requested_user.remove(user_p)
            except Exception as e:
                print(e)

    return redirect(reverse("teacher:lec-detail", kwargs={'pk': lecture.pk}))


def search_lectures(request):
    lecture = request.GET['q']
    print(lecture)
    qs2 = LectrueModel.objects.select_related("teacher").filter(
        Q(lecture_name__icontains=lecture)).filter(college=request.user.user_profile.college)
    qs = LectrueModel.objects.select_related("teacher").filter(Q(lecture_name__icontains=lecture) | Q(
        code__contains=lecture)).filter(college=request.user.user_profile.college)
    print(qs)
    print(qs2)
    return render(request, "teacher/lectures.html", {"objects": qs})


def see_all_sessions(request):
    context = {}
    lecture = request.user.user_profile.lectures.all()
    context['lectures'] = lecture

    return render(request, "teacher/sessions.html", context)


def lecture_accepted_students_from_other_lecture(request, from_pk, to_pk):
    """
    copy other lectures accepted student to other lecture
    """
    user = request.user
    if user.is_teacher and request.POST:
        user_profile = UserProfile.objects.select_related(
            "user").prefect_related("lectures").get(user_id=user.pk)

        lecture = user_profile.lectures.get(pk=to_pk)
        from_copy_lecture = user_profile.lectures.prefetch_related(
            "accepted_user").get(pk=from_pk)

        lecture.accepted_user.set(from_copy_lecture.accepted_user.all())
        lecture.save()
        return redirect(reverse("teacher:lec-detail", kwargs={"pk": to_pk}))
    return redirect("recognizer:login")


def see_all_requests_of_session(request):
    context = {}
    user = request.user
    if user.is_teacher:
        context['requests'] = UserProfile.objects.prefetch_related(
            "change_website_objects", "change_website_objects__requested_users").get(user=user)
    return render(request, "teacher/requests.html", context)


def create_bulk_student(request):
    context = {}
    title = []
    all_the_data = []
    data_of_single = []
    form = StudentBulkForm(request.POST or None, request.FILES or None)
    context["form"] = form

    if request.POST and form.is_valid():
        print(request)
        print(request.FILES)
        print(request.POST)
        file = request.FILES["excel"]
        print(request)
        print(request.FILES)
        print(request.POST)

        dataframe = openpyxl.load_workbook(io.BytesIO(file.read()))

        # Define variable to read sheet
        dataframe1 = dataframe.active

        # Iterate the loop to read the cell values
        print("max rows")
        print(dataframe1.max_row)
        for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                val = col[row].value
                print(row, col, val)
                if row == 0:
                    title.append(val)
                else:
                    data_of_single.append(val)
            all_the_data.append(data_of_single)
            data_of_single = []

        for i in all_the_data:
            if i == []:
                continue
            else:
                create_student(username=i[1], email=i[2], gender=i[3],
                               teacher=request.user.user_profile, enrollment_number=i[4])
                print("sent")

        print(dataframe1)
        print(all_the_data)
        context['form'] = form
        return redirect('teacher:create-bulk-student')
    return render(request, 'teacher/bulk-create.html', context=context)
